"""
Retail Sales Performance Dashboard - Automated Report Generator
=====================================================================
Yeh script cleaned data se ek professional Excel report banata hai.
Isko jab bhi chalao (naya data aane par bhi), 30 second me poora
report ready ho jata hai - jo manually banane me ghanton lagta.

Input:  superstore_cleaned.csv
Output: Retail_Sales_Report.xlsx (4 sheets: Raw Data, Summary,
        Monthly Trend, Category & Region Breakdown)

Isi cheez ko resume me likha hai: "Automated recurring sales
reports, reducing manual reporting time by 60%"
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

# ---------------------------------------------------------
# Config / styling constants
# ---------------------------------------------------------
DATA_PATH = "superstore_cleaned.csv"
OUT_PATH = "Retail_Sales_Report.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=11, bold=True)
KPI_VALUE_FONT = Font(name=FONT_NAME, size=20, bold=True, color="1F4E78")
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
CURRENCY_FMT = '$#,##0'

# ---------------------------------------------------------
# 1. Load cleaned data
# ---------------------------------------------------------
df = pd.read_csv(DATA_PATH)
n_rows = len(df)
print(f"Loaded {n_rows:,} cleaned rows")

wb = Workbook()

# ===========================================================
# SHEET 1: Raw Data (used as the source range for all formulas)
# ===========================================================
ws_raw = wb.active
ws_raw.title = "Raw Data"

cols = list(df.columns)
for c_idx, col_name in enumerate(cols, start=1):
    cell = ws_raw.cell(row=1, column=c_idx, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

for r_idx, row in enumerate(df.itertuples(index=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws_raw.cell(row=r_idx, column=c_idx, value=value).font = BODY_FONT

ws_raw.freeze_panes = "A2"
for c_idx, col_name in enumerate(cols, start=1):
    ws_raw.column_dimensions[get_column_letter(c_idx)].width = max(12, len(col_name) + 2)

last_row = n_rows + 1  # last data row number in Raw Data sheet
category_col = get_column_letter(cols.index("Category") + 1)
region_col = get_column_letter(cols.index("Region") + 1)
sales_col = get_column_letter(cols.index("Sales") + 1)
yearmonth_col = get_column_letter(cols.index("Order Year-Month") + 1)
year_col = get_column_letter(cols.index("Order Year") + 1)
orderid_col = get_column_letter(cols.index("Order ID") + 1)

print(f"Column map -> Category:{category_col} Region:{region_col} Sales:{sales_col} "
      f"YearMonth:{yearmonth_col} Year:{year_col} OrderID:{orderid_col}")

# ===========================================================
# SHEET 2: Summary (KPI cards, formula-driven)
# ===========================================================
ws_sum = wb.create_sheet("Summary")
ws_sum["B2"] = "Retail Sales Performance — Summary"
ws_sum["B2"].font = TITLE_FONT
ws_sum["B3"] = "Auto-generated report — refresh by re-running the script on new data"
ws_sum["B3"].font = Font(name=FONT_NAME, size=9, italic=True, color="808080")

kpis = [
    ("Total Revenue", f"=SUM('Raw Data'!{sales_col}2:{sales_col}{last_row})", CURRENCY_FMT),
    ("Total Orders", f"=SUMPRODUCT(1/COUNTIF('Raw Data'!{orderid_col}2:{orderid_col}{last_row},"
                      f"'Raw Data'!{orderid_col}2:{orderid_col}{last_row}))", '#,##0'),
    ("Total Transactions", f"=COUNTA('Raw Data'!{sales_col}2:{sales_col}{last_row})", '#,##0'),
    ("Avg Transaction Value", f"=AVERAGE('Raw Data'!{sales_col}2:{sales_col}{last_row})", CURRENCY_FMT),
]

row_cursor = 5
for label, formula, fmt in kpis:
    ws_sum.cell(row=row_cursor, column=2, value=label).font = KPI_LABEL_FONT
    val_cell = ws_sum.cell(row=row_cursor + 1, column=2, value=formula)
    val_cell.font = KPI_VALUE_FONT
    val_cell.number_format = fmt
    row_cursor += 3

# Top 3 revenue-driving categories (formula-driven, dynamic)
ws_sum["E4"] = "Top Revenue-Driving Categories"
ws_sum["E4"].font = KPI_LABEL_FONT
ws_sum["E5"] = "Category"
ws_sum["F5"] = "Total Revenue"
ws_sum["G5"] = "% of Total"
for c in ["E5", "F5", "G5"]:
    ws_sum[c].font = HEADER_FONT
    ws_sum[c].fill = HEADER_FILL

categories = sorted(df["Category"].unique())
for i, cat in enumerate(categories, start=6):
    ws_sum.cell(row=i, column=5, value=cat).font = BODY_FONT
    rev_formula = f"=SUMIF('Raw Data'!{category_col}2:{category_col}{last_row},E{i},'Raw Data'!{sales_col}2:{sales_col}{last_row})"
    ws_sum.cell(row=i, column=6, value=rev_formula).number_format = CURRENCY_FMT
    pct_formula = f"=F{i}/$B$6"
    ws_sum.cell(row=i, column=7, value=pct_formula).number_format = "0.0%"

for col, width in [("B", 22), ("C", 14), ("D", 4), ("E", 18), ("F", 16), ("G", 12)]:
    ws_sum.column_dimensions[col].width = width

# ===========================================================
# SHEET 3: Monthly Trend (formula-driven pivot, feeds a line chart)
# ===========================================================
ws_month = wb.create_sheet("Monthly Trend")
ws_month["A1"] = "Order Year-Month"
ws_month["B1"] = "Total Revenue"
for c in ["A1", "B1"]:
    ws_month[c].font = HEADER_FONT
    ws_month[c].fill = HEADER_FILL

months = sorted(df["Order Year-Month"].unique())
for i, ym in enumerate(months, start=2):
    ws_month.cell(row=i, column=1, value=ym).font = BODY_FONT
    formula = f"=SUMIF('Raw Data'!{yearmonth_col}2:{yearmonth_col}{last_row},A{i},'Raw Data'!{sales_col}2:{sales_col}{last_row})"
    ws_month.cell(row=i, column=2, value=formula).number_format = CURRENCY_FMT

ws_month.column_dimensions["A"].width = 18
ws_month.column_dimensions["B"].width = 16

chart1 = LineChart()
chart1.title = "Monthly Revenue Trend"
chart1.style = 2
chart1.y_axis.title = "Revenue"
chart1.x_axis.title = "Month"
data_ref = Reference(ws_month, min_col=2, min_row=1, max_row=len(months) + 1)
cats_ref = Reference(ws_month, min_col=1, min_row=2, max_row=len(months) + 1)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.width = 24
chart1.height = 10
ws_month.add_chart(chart1, "D2")

# ===========================================================
# SHEET 4: Region Breakdown (formula-driven, feeds a bar chart)
# ===========================================================
ws_region = wb.create_sheet("Region Breakdown")
ws_region["A1"] = "Region"
ws_region["B1"] = "Total Revenue"
for c in ["A1", "B1"]:
    ws_region[c].font = HEADER_FONT
    ws_region[c].fill = HEADER_FILL

regions = sorted(df["Region"].unique())
for i, reg in enumerate(regions, start=2):
    ws_region.cell(row=i, column=1, value=reg).font = BODY_FONT
    formula = f"=SUMIF('Raw Data'!{region_col}2:{region_col}{last_row},A{i},'Raw Data'!{sales_col}2:{sales_col}{last_row})"
    ws_region.cell(row=i, column=2, value=formula).number_format = CURRENCY_FMT

ws_region.column_dimensions["A"].width = 14
ws_region.column_dimensions["B"].width = 16

chart2 = BarChart()
chart2.title = "Revenue by Region"
chart2.style = 10
chart2.y_axis.title = "Revenue"
chart2.x_axis.title = "Region"
data_ref2 = Reference(ws_region, min_col=2, min_row=1, max_row=len(regions) + 1)
cats_ref2 = Reference(ws_region, min_col=1, min_row=2, max_row=len(regions) + 1)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
chart2.width = 18
chart2.height = 10
ws_region.add_chart(chart2, "D2")

# ---------------------------------------------------------
# Save
# ---------------------------------------------------------
wb.save(OUT_PATH)
print(f"\nReport saved to: {OUT_PATH}")
